import pandas as pd
import calendar
import random
import unicodedata
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ---------------- FUNCIONES AUXILIARES ----------------
def normalizar(texto):
    return ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')

def asignar_descansos_6x2(horario, persona, estado, dias_mes):
    dias_trabajados = estado.get("contador", 0)
    en_descanso = (estado.get("modo", "T") == "D")
    dias_descanso_restantes = 2 - dias_trabajados if en_descanso else 0

    for d in dias_mes:
        if horario.at[persona, d] in ["D", "F", "V"]:
            continue  # ya tiene un tipo de descanso asignado (prevención de sobrescritura)

        if en_descanso:
            horario.at[persona, d] = "D"
            dias_descanso_restantes -= 1
            if dias_descanso_restantes <= 0:
                en_descanso = False
                dias_trabajados = 0
        else:
            horario.at[persona, d] = None
            dias_trabajados += 1
            if dias_trabajados >= 6:
                en_descanso = True
                dias_descanso_restantes = 2

    # Actualizar estado final
    if en_descanso:
        estado["modo"] = "D"
        estado["contador"] = 2 - dias_descanso_restantes
    else:
        estado["modo"] = "T"
        estado["contador"] = dias_trabajados

def asignar_descansos_5x2(horario, persona, estado, dias_mes):
    dias_trabajados = estado["contador"]
    en_descanso = (estado["modo"] == "D")
    dias_descanso_restantes = 2 - dias_trabajados if en_descanso else 0

    for d in dias_mes:
        if en_descanso:
            horario.at[persona, d] = "D"
            dias_descanso_restantes -= 1
            if dias_descanso_restantes == 0:
                en_descanso = False
                dias_trabajados = 0
        else:
            horario.at[persona, d] = None
            dias_trabajados += 1
            if dias_trabajados == 5:
                en_descanso = True
                dias_descanso_restantes = 2

    estado["modo"] = "D" if en_descanso else "T"
    estado["contador"] = dias_trabajados if not en_descanso else 2 - dias_descanso_restantes

def asignar_descansos_haydee(horario, dias_mes, mes, anio, fase_inicial, modo_inicial, feriados, vacaciones):
    fase = fase_inicial
    dias_trabajados = 0
    dias_descanso = []

    if fase == 1:
        esperando_primer_descanso = (modo_inicial == 'D')
    else:
        esperando_primer_descanso = True

    # ✅ Corrección: si inicia trabajando en fase 1, ya cuenta como trabajado un día
    if fase == 1 and not esperando_primer_descanso:
        dias_trabajados = 1

    for d in dias_mes:
        dia_semana = calendar.day_name[calendar.weekday(anio, mes, d)].lower()

        if fase == 1:
            if esperando_primer_descanso:
                if dia_semana == "monday":
                    dias_descanso.append(d)
                elif dia_semana == "tuesday":
                    dias_descanso.append(d)
                    esperando_primer_descanso = False
                    dias_trabajados = 0
                    fase = 2
                continue

            if dias_trabajados < 7:
                dias_trabajados += 1
            else:
                if dia_semana == "monday":
                    dias_descanso.append(d)
                elif dia_semana == "tuesday":
                    dias_descanso.append(d)
                    dias_trabajados = 0
                    fase = 2
                else:
                    dias_trabajados += 1

        elif fase == 2:
            if esperando_primer_descanso:
                if dia_semana == "saturday":
                    dias_descanso.append(d)
                elif dia_semana == "sunday":
                    dias_descanso.append(d)
                    esperando_primer_descanso = False
                    dias_trabajados = 0
                    fase = 1
                continue

            if dias_trabajados < 3:
                dias_trabajados += 1
            else:
                if dia_semana == "saturday":
                    dias_descanso.append(d)
                elif dia_semana == "sunday":
                    dias_descanso.append(d)
                    dias_trabajados = 0
                    fase = 1
                else:
                    dias_trabajados += 1

    for d in dias_descanso:
        horario.at["Haydee Fernandez", d] = "D"
        print(f"\u2705 Descanso asignado a Haydee el día {d} ({calendar.day_name[calendar.weekday(anio, mes, d)]})")

def asignar_roles(d):
    # Alternar orden para que Silvia no siempre tenga Uso/DAAF
    roles_prioritarios = ["Uso/DAAF", "P/NI"] if d % 2 == 0 else ["P/NI", "Uso/DAAF"]
    rol_opcional = "OA"

    def esta_disponible(persona):
        return horario.at[persona, d] not in ["D", "F", "V", "ADM"]

    # Listas de asistentes disponibles
    disponibles = [a for a in asistentes_52 + asistentes_62 if esta_disponible(a)]
    disponibles_sin_haydee = [a for a in disponibles if a != "Haydee Fernandez"]

    # 1. Asignar Uso/DAAF y P/NI
    for rol in roles_prioritarios:
        if rol == "P/NI":
            candidatos = disponibles_sin_haydee
        else:
            candidatos = disponibles
        candidatos = [a for a in candidatos if esta_disponible(a)]
        random.shuffle(candidatos)
        if candidatos:
            seleccionado = candidatos.pop(0)
            horario.at[seleccionado, d] = rol
            if seleccionado in disponibles:
                disponibles.remove(seleccionado)
            if seleccionado in disponibles_sin_haydee:
                disponibles_sin_haydee.remove(seleccionado)

    # 2. Asignar OA si queda alguien libre
    candidatos_oa = [a for a in disponibles if esta_disponible(a)]
    random.shuffle(candidatos_oa)
    if candidatos_oa:
        seleccionado = candidatos_oa[0]
        horario.at[seleccionado, d] = rol_opcional
        disponibles.remove(seleccionado)

    # 3. Rellenar asistentes restantes con cualquier rol permitido
    for a in disponibles:
        if esta_disponible(a):
            if a == "Haydee Fernandez":
                opciones = ["Uso/DAAF", "OA"]  # sin P/NI
            else:
                opciones = ["Uso/DAAF", "P/NI", "OA"]
            horario.at[a, d] = random.choice(opciones)

    # 4. Asignar asesores
    libres_ases = [a for a in asesores if esta_disponible(a)]

    # TT obligatorio
    if libres_ases:
        candidato_tt = min(libres_ases, key=lambda x: contador_tt[x])
        horario.at[candidato_tt, d] = "TT"
        contador_tt[candidato_tt] += 1
        libres_ases.remove(candidato_tt)

    # N si Jackeline no está
    if horario.at["Jackeline Tapia", d] in ["D", "F", "V"] and libres_ases:
        horario.at[libres_ases.pop(), d] = "N"

    # OA si ningún asistente lo tomó
    hay_oa = any(horario.loc[asistentes_52 + asistentes_62 + asesores, d] == "OA")
    if not hay_oa and libres_ases:
        for a in libres_ases:
            if esta_disponible(a):
                horario.at[a, d] = "OA"
                libres_ases.remove(a)
                break

    # Rellenar asesores restantes con TD
    for a in libres_ases:
        if esta_disponible(a):
            horario.at[a, d] = "TD"

def asignar_feriados_compensados(horario, feriados, personas, feriado_largo):
    roles_asesores = ["TT", "TD", "N"]
    roles_asistentes = ["P/NI", "Uso/DAAF", "OA"]

    feriados_trabajados = {p: 0 for p in personas}

    # 1. Contar feriados realmente trabajados
    for dia in feriados:
        for p in personas:
            if horario.at[p, dia] not in ["D", "F", "V"]:
                feriados_trabajados[p] += 1

    # 2. Buscar días adecuados para colocar 'F' compensatorio
    def puede_reemplazar(p, d):
        if horario.at[p, d] in ["D", "F", "V", "ADM"]:
            return False
        rol = horario.at[p, d]

        # Validación de cobertura
        if p in asesores:
            total = sum(horario.loc[asesores, d] == rol)
            if rol == "N" and horario.at["Jackeline Tapia", d] in ["D", "F", "V"]:
                return total > 1
            return total > 1 if rol in roles_asesores else False
        else:
            total = sum(horario.loc[asistentes_52 + asistentes_62, d] == rol)
            return total > 1 if rol in roles_asistentes else False

    for p, cantidad in feriados_trabajados.items():
        if cantidad == 0:
            continue

        # Generar candidatos disponibles
        disponibles = [d for d in horario.columns if d not in feriados and horario.at[p, d] not in ["D", "F", "V", "ADM"]]
        
        # 3. Si es feriado largo, prioriza días junto a descansos (DDFF o FFDD)
        preferidos = []
        if feriado_largo:
            for d in disponibles:
                vecinos = [d - 1, d + 1]
                if any(horario.at[p, v] == "D" for v in vecinos if v in horario.columns):
                    preferidos.append(d)

        # Evitar primeros días del mes si no es necesario
        tardíos = [d for d in disponibles if d > 3]
        tempranos = [d for d in disponibles if d <= 3]

        # Orden final de candidatos
        candidatos = list(dict.fromkeys(preferidos + tardíos + tempranos))

        # Asignar los 'F'
        for d in candidatos:
            if puede_reemplazar(p, d):
                horario.at[p, d] = "F"
                cantidad -= 1
                if cantidad == 0:
                    break

        # 🚨 Mensaje si no se pudo compensar completamente
        if cantidad > 0:
            print(f"⚠️ No se pudo asignar todos los F para {p}, faltaron {cantidad}")



# NUEVA FUNCIÓN: ASIGNAR DÍAS ADMINISTRATIVOS
def asignar_adm(horario, dias_mes):
    esenciales_asist = ["P/NI", "Uso/DAAF", "OA"]
    esenciales_ases = ["TT", "TD"]
    grupo_a = ["Franklin Córdova", "Fredey Flores", "Javier Cano", "Julio Rodriguez", "Carlos Ramos",
               "Patrick Romero", "Luis Arancibia", "Haydee Fernandez"]
    grupo_b = ["Barbara Severino", "Guisela Meneses"]
    grupo_c = ["Silvia Hernandez"]

    conteo_adm = {p: 0 for p in grupo_a + grupo_b + grupo_c}

    # Guisela y Barbara: ideal 3 ADM, no pegados
    for p in grupo_b:
        dias_posibles = [d for d in dias_mes if horario.at[p, d] not in ["D", "F", "V", "ADM"]]
        dispersos = sorted([d for d in dias_posibles if 4 < d < len(dias_mes) - 3])
        seleccionados = []
        for d in dispersos:
            if any(abs(d - s) < 5 for s in seleccionados):
                continue
            rol = horario.at[p, d]
            if rol in esenciales_asist and sum(horario.loc[:, d] == rol) > 1:
                horario.at[p, d] = "ADM"
                conteo_adm[p] += 1
                seleccionados.append(d)
            if conteo_adm[p] >= 3:
                break

    # Grupo A: asegurar 1 ADM si posible
    for p in grupo_a:
        dias_posibles = [d for d in dias_mes if horario.at[p, d] not in ["D", "F", "V", "ADM"]]
        random.shuffle(dias_posibles)
        for d in dias_posibles:
            rol = horario.at[p, d]
            if p in asesores and rol in esenciales_ases and sum(horario.loc[asesores, d] == rol) > 1:
                horario.at[p, d] = "ADM"
                conteo_adm[p] += 1
                break
            elif p not in asesores and rol in esenciales_asist and sum(horario.loc[:, d] == rol) > 1:
                horario.at[p, d] = "ADM"
                conteo_adm[p] += 1
                break

    # Silvia: intentar obtener la mayor cantidad de ADM posible
    p = "Silvia Hernandez"
    dias_posibles = [d for d in dias_mes if horario.at[p, d] not in ["D", "F", "V", "ADM"]]

    for d in dias_posibles:
        rol_silvia = horario.at[p, d]

        # Caso 1: Silvia no es la única con ese rol → ADM directa
        if rol_silvia in esenciales_asist and sum(horario.loc[asistentes_52 + asistentes_62, d] == rol_silvia) > 1:
            horario.at[p, d] = "ADM"
            conteo_adm[p] += 1
            continue

        # Caso 2: Silvia es única, pero puede ceder su rol a otro con rol repetido
        for otro in asistentes_52 + asistentes_62:
            if otro == p:
                continue
            rol_otro = horario.at[otro, d]
            if rol_otro != rol_silvia and rol_otro in esenciales_asist:
                if sum(horario.loc[asistentes_52 + asistentes_62, d] == rol_otro) > 1:
                    # Intercambiar roles
                    horario.at[otro, d] = rol_silvia
                    horario.at[p, d] = "ADM"
                    conteo_adm[p] += 1
                    break

    print("✔️ ADM asignado correctamente.")

def ajustar_td_exceso(horario, asesores):
    for d in horario.columns:
        td_en_dia = [a for a in asesores if horario.at[a, d] == "TD"]
        if len(td_en_dia) >= 4:
            # Cambiar uno de los TD a OA (el último, por ejemplo, para no alterar los primeros)
            persona_a_cambiar = td_en_dia[-1]
            horario.at[persona_a_cambiar, d] = "OA"
            print(f"🔁 {persona_a_cambiar} cambiado de TD a OA en el día {d}")

# ---------------- CONFIGURACIÓN ----------------
asistentes_52 = ["Silvia Hernandez", "Guisela Meneses", "Barbara Severino"]
asistentes_62 = ["Haydee Fernandez", "Patrick Romero", "Luis Arancibia"]
asesores = ["Franklin Córdova", "Fredey Flores", "Javier Cano", "Julio Rodriguez", "Carlos Ramos"]
call = ["Jackeline Tapia"]
personas = asesores + asistentes_52 + asistentes_62 + call

# ---------------- INPUTS ----------------
mes_num = int(input("Mes (1-12): "))
anio = int(input("Año: "))
num_dias = calendar.monthrange(anio, mes_num)[1]
dias_mes = list(range(1, num_dias + 1))

dia_descanso_call = normalizar(input("¿Qué día descansa Jackeline Tapia?: ").strip().lower())
dias_estandar = {"lunes": "monday", "martes": "tuesday", "miercoles": "wednesday", "jueves": "thursday", "viernes": "friday", "sabado": "saturday", "domingo": "sunday"}
dia_target = dias_estandar.get(dia_descanso_call)
if not dia_target:
    print("Día no válido."); exit()

num_feriados = int(input("¿Cuántos feriados hay este mes?: "))
feriados = [int(input(f"Ingrese el día del feriado #{i+1}: ")) for i in range(num_feriados)]
feriado_largo = input("¿Hay feriado largo? (s/n): ").strip().lower() == 's'

vacaciones_jackeline = []
if input("Jackeline tiene vacaciones este mes? (s/n): ").strip().lower() == 's':
    num_vacaciones = int(input("Cuántos días?: "))
    vacaciones_jackeline = [int(input(f"Día de vacaciones #{i+1}: ")) for i in range(num_vacaciones)]

estado_descanso = {}
contador_tt = {a: 0 for a in asesores}

for p in asesores + ["Patrick Romero", "Luis Arancibia"]:
    dias_faltan = int(input(f"¿Cuántos días le faltan a {p} para su descanso (6x2)? (0 si ya está en descanso): "))
    if dias_faltan == 0:
        dia_actual = int(input("Está en día 1 o 2 de descanso?: "))
        estado_descanso[p] = {"modo": "D", "contador": 1 if dia_actual == 2 else 2}
    else:
        estado_descanso[p] = {"modo": "T", "contador": 6 - dias_faltan}

for p in asistentes_52:
    dias_faltan = int(input(f"¿Cuántos días le faltan a {p} para su descanso (5x2)? (0 si ya está en descanso): "))
    if dias_faltan == 0:
        dia_actual = int(input("Está en día 1 o 2 de descanso?: "))
        estado_descanso[p] = {"modo": "D", "contador": 1 if dia_actual == 2 else 2}
    else:
        estado_descanso[p] = {"modo": "T", "contador": 5 - dias_faltan}

fase_haydee = int(input("¿En qué fase empieza Haydee? (1 = trabaja L-D, descansa L-M / 2 = trabaja M-V, descansa S-D): "))
if fase_haydee == 1:
    modo_haydee = input("¿Haydee inicia descansando o trabajando? (D/T): ").strip().upper()
    if modo_haydee not in ["D", "T"]:
        print("Modo inválido. Usa 'D' para descanso o 'T' para trabajo.")
        exit()
else:
    modo_haydee = "T"  # No afecta, pero se pasa igual para compatibilidad

# ---------------- GENERAR HORARIO ----------------
horario = pd.DataFrame(index=personas, columns=dias_mes)

for d in dias_mes:
    dia_actual = calendar.day_name[calendar.weekday(anio, mes_num, d)].lower()
    if d in vacaciones_jackeline:
        horario.at["Jackeline Tapia", d] = "V"
    elif d in feriados:
        horario.at["Jackeline Tapia", d] = "F"
    elif dia_actual == dia_target:
        horario.at["Jackeline Tapia", d] = "D"
    else:
        horario.at["Jackeline Tapia", d] = "N"

for p in asesores + ["Patrick Romero", "Luis Arancibia"]:
    asignar_descansos_6x2(horario, p, estado_descanso[p], dias_mes)

for p in asistentes_52:
    asignar_descansos_5x2(horario, p, estado_descanso[p], dias_mes)

asignar_descansos_haydee(horario, dias_mes, mes_num, anio, fase_haydee, modo_haydee, feriados, vacaciones_jackeline)

for d in dias_mes:
    asignar_roles(d)

asignar_feriados_compensados(horario, feriados, asesores + asistentes_52 + asistentes_62, feriado_largo)

asignar_adm(horario, dias_mes)

ajustar_td_exceso(horario, asesores)

# ---------------- EXPORTAR CON FORMATO ----------------
# from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

archivo = f"Horario_Completo_{calendar.month_name[mes_num]}_{anio}.xlsx"

with pd.ExcelWriter(archivo, engine="openpyxl") as writer:
    horario.to_excel(writer, startrow=1)
    ws = writer.sheets['Sheet1']

    # Estilos generales
    gris = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    negrita_arial_14 = Font(name='Arial', size=14, bold=True)
    negrita_arial_12 = Font(name='Arial', size=12, bold=True)
    negrita_arial_10 = Font(name='Arial', size=10, bold=True)
    fuente_contenido = Font(name='Times New Roman', size=7)
    fuente_personal = Font(name='Calibri', size=13, color='093D93')

    # Formatos por contenido
    colores_celdas = {
        "TT": PatternFill(start_color="BC8E03", end_color="BC8E03", fill_type="solid"),
        "N": PatternFill(start_color="7AD694", end_color="7AD694", fill_type="solid"),
        "Uso/DAAF": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        "P/NI": PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid"),
        "OA": PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid"),
        "D": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        "F": PatternFill(start_color="002060", end_color="002060", fill_type="solid"),
        "V": PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid"),
        "ADM": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    }

    color_fuente_f = Font(name='Times New Roman', size=7, color="FFFFFF")

    # Título del mes
    ws.cell(row=1, column=1, value=calendar.month_name[mes_num])
    ws.cell(row=1, column=1).font = negrita_arial_14
    ws.cell(row=1, column=1).fill = gris
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

    # Subtítulo
    ws.cell(row=2, column=1, value="Personal de SIS")
    ws.cell(row=2, column=1).font = negrita_arial_12
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')

    # Encabezados de días (fila 1, columnas 2+)
    dias_cortos = ["L", "M", "M", "J", "V", "S", "D"]
    for col, d in enumerate(dias_mes, start=2):
        dia_semana = calendar.weekday(anio, mes_num, d)
        letra = dias_cortos[dia_semana]
        celda = ws.cell(row=1, column=col, value=letra)
        celda.font = negrita_arial_10
        celda.fill = rojo if dia_semana in [5, 6] else gris
        celda.alignment = Alignment(horizontal='center', vertical='center')

    # Centrado de fila 2 (números de días)
    max_col = ws.max_column
    for col in range(2, max_col + 1):
        celda = ws.cell(row=2, column=col)
        celda.alignment = Alignment(horizontal='center', vertical='center')

    # Formato y colores para los datos
    max_row = ws.max_row
    for row in range(3, max_row + 1):
        # Columna 1 (nombres)
        celda_nombre = ws.cell(row=row, column=1)
        celda_nombre.font = fuente_personal
        celda_nombre.alignment = Alignment(horizontal='left', vertical='center')

        # Celdas de datos
        for col in range(2, max_col + 1):
            celda = ws.cell(row=row, column=col)
            valor = str(celda.value).strip()
            celda.font = fuente_contenido
            celda.alignment = Alignment(horizontal='center', vertical='center')
            if valor in colores_celdas:
                celda.fill = colores_celdas[valor]
                if valor == "F":
                    celda.font = color_fuente_f  # blanco sobre azul oscuro

    # Bordes exteriores medium, interiores thin + fila 2 y col 1 destacados
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            bottom_style = 'medium' if row == max_row or row == 2 else 'thin'
            right_style = 'medium' if col == max_col or col == 1 else 'thin'
            ws.cell(row=row, column=col).border = Border(
                left=Side(style='medium' if col == 1 else 'thin'),
                right=Side(style=right_style),
                top=Side(style='medium' if row == 1 else 'thin'),
                bottom=Side(style=bottom_style)
            )

    # Ajuste automático de ancho de columna
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

print(f"\n✅ Horario COMPLETO FINAL generado con formato: {archivo}")