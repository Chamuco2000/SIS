import tkinter as tk
from tkinter import messagebox, simpledialog
import calendar
import pandas as pd
import random
import unicodedata
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Variables globales
mes_num = anio = 0
feriados = []
vacaciones_jackeline = []
dia_descanso_call = ""
fase_haydee = 1
estado_descanso = {}
contador_tt = {}
horario = None

# Datos fijos
asistentes_52 = ["Silvia Hernandez", "Guisela Meneses", "Barbara Severino"]
asistentes_62 = ["Haydee Fernandez", "Patrick Romero", "Luis Arancibia"]
asesores = ["Franklin Córdova", "Fredey Flores", "Javier Cano", "Julio Rodriguez", "Carlos Ramos"]
call = ["Jackeline Tapia"]
personas = asesores + asistentes_52 + asistentes_62 + call

def normalizar(texto):
    return ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')

def asignar_descansos_6x2(horario, persona, estado, dias_mes):
    dias_trabajados = estado["contador"]
    en_descanso = (estado["modo"] == "D")
    dias_descanso_restantes = 2 - dias_trabajados if en_descanso else 0

    for d in dias_mes:
        if en_descanso:
            horario.at[persona, d] = "D"
            dias_descanso_restantes -= 1
            if dias_descanso_restantes == 0:
                en_descanso = False
                dias_trabajados = 0
        else:
            horario.at[persona, d] = None
            dias_trabajados += 1
            if dias_trabajados == 6:
                en_descanso = True
                dias_descanso_restantes = 2

    estado["modo"] = "D" if en_descanso else "T"
    estado["contador"] = dias_trabajados if not en_descanso else 2 - dias_descanso_restantes

def asignar_descansos_5x2(horario, persona, estado, dias_mes):
    dias_trabajados = estado["contador"]
    en_descanso = (estado["modo"] == "D")
    dias_descanso_restantes = 2 - dias_trabajados if en_descanso else 0

    for d in dias_mes:
        if en_descanso:
            horario.at[persona, d] = "D"
            dias_descanso_restantes -= 1
            if dias_descanso_restantes == 0:
                en_descanso = False
                dias_trabajados = 0
        else:
            horario.at[persona, d] = None
            dias_trabajados += 1
            if dias_trabajados == 5:
                en_descanso = True
                dias_descanso_restantes = 2

    estado["modo"] = "D" if en_descanso else "T"
    estado["contador"] = dias_trabajados if not en_descanso else 2 - dias_descanso_restantes

def asignar_descansos_haydee(horario, dias_mes, mes, anio, fase_inicial, feriados, vacaciones):
    fase = fase_inicial
    dias_trabajados = 0
    esperando_primer_descanso = True
    dias_descanso = []

    for d in dias_mes:
        dia_semana = calendar.day_name[calendar.weekday(anio, mes, d)].lower()
        es_feriado = d in feriados or d in vacaciones

        if fase == 1:
            if esperando_primer_descanso and dia_semana in ["monday", "tuesday"]:
                if not es_feriado:
                    dias_descanso.append(d)
                if dia_semana == "tuesday":
                    esperando_primer_descanso = False
                    dias_trabajados = 0
                    fase = 2
            else:
                if dias_trabajados < 7:
                    dias_trabajados += 1
                else:
                    if not es_feriado and dia_semana in ["monday", "tuesday"]:
                        dias_descanso.append(d)
                    if dia_semana == "tuesday":
                        dias_trabajados = 0
                        fase = 2
                    else:
                        dias_trabajados += 1

        elif fase == 2:
            if esperando_primer_descanso and dia_semana in ["saturday", "sunday"]:
                if not es_feriado:
                    dias_descanso.append(d)
                if dia_semana == "sunday":
                    esperando_primer_descanso = False
                    dias_trabajados = 0
                    fase = 1
            else:
                if dias_trabajados < 3:
                    dias_trabajados += 1
                else:
                    if not es_feriado and dia_semana in ["saturday", "sunday"]:
                        dias_descanso.append(d)
                    if dia_semana == "sunday":
                        dias_trabajados = 0
                        fase = 1
                    else:
                        dias_trabajados += 1

    for d in dias_descanso:
        horario.at["Haydee Fernandez", d] = "D"

def exportar_y_abrir_archivo(horario, mes_num, anio):
    archivo = f"horario_completo_{calendar.month_name[mes_num]}_{anio}.xlsx"

    with pd.ExcelWriter(archivo, engine="openpyxl") as writer:
        horario.to_excel(writer, startrow=1)
        ws = writer.sheets['Sheet1']

        # Estilos
        gris = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        negrita_arial_14 = Font(name='Arial', size=14, bold=True)
        negrita_arial_12 = Font(name='Arial', size=12, bold=True)
        negrita_arial_10 = Font(name='Arial', size=10, bold=True)

        dias_mes = list(horario.columns)
        dias_cortos = ["L", "M", "M", "J", "V", "S", "D"]

        # Nombre mes
        ws.cell(row=1, column=1, value=calendar.month_name[mes_num])
        ws.cell(row=1, column=1).font = negrita_arial_14
        ws.cell(row=1, column=1).fill = gris
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

        # Subtitulo
        ws.cell(row=2, column=1, value="Personal de SIS")
        ws.cell(row=2, column=1).font = negrita_arial_12
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')

        # Letras
        for col, d in enumerate(dias_mes, start=2):
            dia_semana = calendar.weekday(anio, mes_num, d)
            letra = dias_cortos[dia_semana]
            celda = ws.cell(row=1, column=col, value=letra)
            celda.font = negrita_arial_10
            celda.fill = rojo if dia_semana in [5,6] else gris
            celda.alignment = Alignment(horizontal='center', vertical='center')

    os.startfile(archivo)
