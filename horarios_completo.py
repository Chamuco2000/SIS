# horarios_completo.py
import pandas as pd
import calendar
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from utils_horarios import normalizar
import random

# Aquí debes copiar todas tus funciones: asignar_descansos_6x2, asignar_descansos_5x2, etc.
# Me ahorro volverlas a copiar aquí ya que las conservas y solo cambiaremos el main
# Asegúrate de que este archivo NO incluya inputs, sino solo funciones

# CONFIGURACIÓN GENERAL
asistentes_52 = ["Silvia Hernandez", "Guisela Meneses", "Barbara Severino"]
asistentes_62 = ["Haydee Fernandez", "Patrick Romero", "Luis Arancibia"]
asesores = ["Franklin Córdova", "Fredey Flores", "Javier Cano", "Julio Rodriguez", "Carlos Ramos"]
call = ["Jackeline Tapia"]
personas = asesores + asistentes_52 + asistentes_62 + call


def generar_horario(mes_num, anio, dia_descanso_call, feriados, feriado_largo, vacaciones_jackeline,
                    estado_descanso, fase_haydee, modo_haydee):
    num_dias = calendar.monthrange(anio, mes_num)[1]
    dias_mes = list(range(1, num_dias + 1))

    global horario, contador_tt
    horario = pd.DataFrame(index=personas, columns=dias_mes)
    contador_tt = {a: 0 for a in asesores}

    dias_estandar = {
        "lunes": "monday", "martes": "tuesday", "miercoles": "wednesday",
        "jueves": "thursday", "viernes": "friday", "sabado": "saturday", "domingo": "sunday"
    }
    dia_target = dias_estandar.get(normalizar(dia_descanso_call.strip().lower()))

    for d in dias_mes:
        dia_actual = calendar.day_name[calendar.weekday(anio, mes_num, d)].lower()
        if d in vacaciones_jackeline:
            horario.at["Jackeline Tapia", d] = "V"
        elif d in feriados:
            horario.at["Jackeline Tapia", d] = "F"
        elif dia_actual == dia_target:
            horario.at["Jackeline Tapia", d] = "D"
        else:
            horario.at["Jackeline Tapia", d] = "N"

    for p in asesores + ["Patrick Romero", "Luis Arancibia"]:
        asignar_descansos_6x2(horario, p, estado_descanso[p], dias_mes)

    for p in asistentes_52:
        asignar_descansos_5x2(horario, p, estado_descanso[p], dias_mes)

    asignar_descansos_haydee(horario, dias_mes, mes_num, anio, fase_haydee, modo_haydee, feriados, vacaciones_jackeline)

    for d in dias_mes:
        asignar_roles(d)

    asignar_feriados_compensados(horario, feriados, asesores + asistentes_52 + asistentes_62, feriado_largo)
    asignar_adm(horario, dias_mes)
    ajustar_td_exceso(horario, asesores)

    return horario


def exportar_horario(horario, mes_num, anio):
    archivo = f"Horario_Completo_{calendar.month_name[mes_num]}_{anio}.xlsx"

    with pd.ExcelWriter(archivo, engine="openpyxl") as writer:
        horario.to_excel(writer, startrow=1)
        ws = writer.sheets['Sheet1']

        # Estilos y formato (mismos del original, resumidos aquí para claridad)
        gris = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        negrita_arial_14 = Font(name='Arial', size=14, bold=True)
        negrita_arial_12 = Font(name='Arial', size=12, bold=True)
        negrita_arial_10 = Font(name='Arial', size=10, bold=True)
        fuente_contenido = Font(name='Times New Roman', size=7)
        fuente_personal = Font(name='Calibri', size=13, color='093D93')

        colores_celdas = {
            "TT": PatternFill(start_color="BC8E03", end_color="BC8E03", fill_type="solid"),
            "N": PatternFill(start_color="7AD694", end_color="7AD694", fill_type="solid"),
            "Uso/DAAF": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
            "P/NI": PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid"),
            "OA": PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid"),
            "D": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            "F": PatternFill(start_color="002060", end_color="002060", fill_type="solid"),
            "V": PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid"),
            "ADM": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        }
        color_fuente_f = Font(name='Times New Roman', size=7, color="FFFFFF")

        ws.cell(row=1, column=1, value=calendar.month_name[mes_num])
        ws.cell(row=1, column=1).font = negrita_arial_14
        ws.cell(row=1, column=1).fill = gris
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=1, value="Personal de SIS")
        ws.cell(row=2, column=1).font = negrita_arial_12
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')

        dias_cortos = ["L", "M", "M", "J", "V", "S", "D"]
        dias_mes = list(horario.columns)
        for col, d in enumerate(dias_mes, start=2):
            dia_semana = calendar.weekday(anio, mes_num, d)
            letra = dias_cortos[dia_semana]
            celda = ws.cell(row=1, column=col, value=letra)
            celda.font = negrita_arial_10
            celda.fill = rojo if dia_semana in [5, 6] else gris
            celda.alignment = Alignment(horizontal='center', vertical='center')

        max_col = ws.max_column
        for col in range(2, max_col + 1):
            ws.cell(row=2, column=col).alignment = Alignment(horizontal='center', vertical='center')

        max_row = ws.max_row
        for row in range(3, max_row + 1):
            celda_nombre = ws.cell(row=row, column=1)
            celda_nombre.font = fuente_personal
            celda_nombre.alignment = Alignment(horizontal='left', vertical='center')

            for col in range(2, max_col + 1):
                celda = ws.cell(row=row, column=col)
                valor = str(celda.value).strip()
                celda.font = fuente_contenido
                celda.alignment = Alignment(horizontal='center', vertical='center')
                if valor in colores_celdas:
                    celda.fill = colores_celdas[valor]
                    if valor == "F":
                        celda.font = color_fuente_f

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                bottom_style = 'medium' if row == max_row or row == 2 else 'thin'
                right_style = 'medium' if col == max_col or col == 1 else 'thin'
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style='medium' if col == 1 else 'thin'),
                    right=Side(style=right_style),
                    top=Side(style='medium' if row == 1 else 'thin'),
                    bottom=Side(style=bottom_style)
                )

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

    print(f"\n✅ Horario COMPLETO FINAL generado con formato: {archivo}")

